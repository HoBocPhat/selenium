# from contextlib import nullcontext
import openpyxl
import time
# import excel
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.common.exceptions import NoSuchElementException
# import selenium
# print (selenium.__version__)

driver = webdriver.Chrome('.\chromedriver.exe')

driver.maximize_window()

driver.get("http://127.0.0.1:8000/")


path = ".\login.xlsx"
testcase = openpyxl.load_workbook(path)
sheet = testcase.active
rows = sheet.max_row
cols = sheet.max_column
# print(rows)
# print(cols)

def clear_text(element):
    length = len(element.get_attribute('value'))
    element.send_keys(length * Keys.BACKSPACE)

def check_exists_by_xpath(xpath):
    try:
        driver.find_element_by_xpath(xpath)
    except NoSuchElementException:
        return False
    return True 
driver.find_element_by_xpath('/html/body/div[1]/div[3]/p/a[1]').click()

testcase = rows - 1
tested = 0
succeed = 0
failed = 0
i=0
for r in range(2,rows+1):
    i = + 1
    username_data = sheet.cell(row=r,column=1).value
    password_data = sheet.cell(row=r,column=2).value
    # print(username, password)


#Log in
#Điền tên đăng nhập, mật khẩu       
    username = driver.find_element_by_xpath('//*[@id="id_username"]')
    
    clear_text(username)
    if (username_data != None): 
        driver.find_element_by_xpath('//*[@id="id_username"]').send_keys(username_data)
    
    password = driver.find_element_by_xpath('//*[@id="id_password"]')
    clear_text(password)
    
    if (password_data != None):
        driver.find_element_by_xpath('//*[@id="id_password"]').send_keys(password_data)
    
    driver.find_element_by_xpath('/html/body/div[3]/div/div/form/button').click() 
    time.sleep(3)
    
    if(check_exists_by_xpath('/html/body/div[1]/div[3]/p/a[1]') == False):
        driver.find_element_by_xpath('/html/body/div[1]/div[3]/div/span/a').click()
        tested +=1
        print("Test case ", i ," return success")
        succeed +=1
        time.sleep(3)
        driver.find_element_by_xpath('/html/body/div[1]/div[3]/p/a[1]').click()
    
    else:
        tested +=1
        print("Test case ", i ," return fail")
        failed +=1 
        driver.find_element_by_xpath('/html/body/div[1]/div[3]/p/a[1]').click()
    
driver.close()

print("Testing report")
print("Testcases: ",testcase)
print("Untested: ", testcase - tested)
print("Tested: ", tested)
print("Testcases return success: ", succeed)
print("Testcases return fail: ", failed)