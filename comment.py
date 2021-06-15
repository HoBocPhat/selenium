import openpyxl
import time
# import excel
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
driver = webdriver.Chrome('.\chromedriver.exe')

driver.maximize_window()

driver.get("http://127.0.0.1:8000/")

path = ".\comment.xlsx"
testcase = openpyxl.load_workbook(path)
sheet = testcase.active
rows = sheet.max_row
testcase = rows - 1
tested = 0
succeed = 0
failed = 0
i = 0
def clear_text(element):
    length = len(element.get_attribute('value'))
    element.send_keys(length * Keys.BACKSPACE)
def check_exists_by_xpath(xpath):
    try:
        driver.find_element_by_xpath(xpath)
    except NoSuchElementException:
        return False
    return True 
driver.find_element_by_xpath('/html/body/div[1]/div[3]/p/a[1]').click()
driver.find_element_by_xpath('//*[@id="id_username"]').send_keys("1")
driver.find_element_by_xpath('//*[@id="id_password"]').send_keys("Ngotuongvy123")
driver.find_element_by_xpath('/html/body/div[3]/div/div/form/button').click() 
driver.find_element_by_xpath('//*[@id="carousel-example1"]/div[1]/div/div/a[1]/div/div/div[1]/picture/img').click()
lastcomment = '/html/body/div[3]/div[4]/div/div/div/div/div[3]/div[3]/div[3]'
for r in range(2,rows+1):
    i +=1
    comment_data = sheet.cell(row=r,column=1).value
   
    comment = driver.find_element_by_xpath('//*[@id="id_content"]')
    clear_text(comment)
    if comment_data != None:
        driver.find_element_by_xpath('//*[@id="id_content"]').send_keys(comment_data)
    driver.find_element_by_xpath('/html/body/div[3]/div[3]/div/div/div/form/button').click()
    
    newcomment = lastcomment + '/div[3]'
    if(check_exists_by_xpath(newcomment) == True):
        tested +=1
        print("Testcase ", i , "return success")
        lastcomment = newcomment
        succeed +=1
    else:
        tested +=1
        print("Testcase ", i, "return fail")
        failed +=1
    time.sleep(3)
driver.close()

print("Testing report")
print("Testcases: ",testcase)
print("Untested: ", testcase - tested)
print("Tested: ", tested)
print("Testcases return success: ", succeed)
print("Testcases return fail: ", failed)